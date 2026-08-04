from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from unittest.mock import patch

import openpyxl
import pytest

from src.data import DataSourceError, FailedRow, Header, InventoryRow, LandedCostRow
from src.data.datarows import InventoryDTO, LandedCostDTO
from src.data.excel.reader import ExcelFileReader
from src.adapters import allocate_landed_costs, build_inventory, give_reader


class _FakeReader:
    """Minimal stand-in for excel.Reader[T] -- build_inventory/allocate_landed_costs
    only ever call .readline(), so a real workbook isn't needed to exercise them."""

    def __init__(self, rows):
        self._rows = rows

    def readline(self):
        return iter(self._rows)


def _inventory_row(sku: str, qty: int) -> InventoryRow:
    dto = InventoryDTO()
    dto.sku = sku
    dto.base_sku = sku
    dto.inventory = Decimal(qty)
    return InventoryRow(dto)


def _landed_cost_row(sku: str, qty: int, unit_cost: int) -> LandedCostRow:
    dto = LandedCostDTO()
    dto.sku = sku
    dto.qty = Decimal(qty)
    dto.unit_cost = Decimal(unit_cost)
    dto.date = datetime(2024, 1, 1)
    return LandedCostRow(dto)


class TestGiveReader:
    def test_raises_typeerror_when_file_is_not_xlsx(self):
        with pytest.raises(TypeError):
            give_reader(
                file_path="inventory.csv",
                sheet_name="Sheet1",
                header=Header.inventory_row(sku=0, base_sku=1, inventory=2),
                return_type=InventoryRow,
            )

    def test_wires_reader_with_requested_header_and_return_type(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Inventory"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        real_wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

        header = Header.inventory_row(sku=0, base_sku=1, inventory=2)
        with patch("src.data.excel.reader.xl.load_workbook", return_value=real_wb):
            reader = give_reader(
                file_path="inventory.xlsx",
                sheet_name="Inventory",
                header=header,
                return_type=InventoryRow,
            )

        assert isinstance(reader, ExcelFileReader)
        assert reader.rt is InventoryRow
        assert reader.header is header


class TestBuildInventory:
    def test_aggregates_duplicate_skus(self):
        rows = [_inventory_row("sku-a", 10), _inventory_row("sku-a", 5)]

        inventory, failed = build_inventory(_FakeReader(rows))

        assert failed == []
        assert inventory["sku-a"].qty == Decimal(15)
        # unallocated is reset to the newly-merged total, per build_inventory's merge step
        assert inventory["sku-a"].unallocated == Decimal(15)

    def test_collects_failed_rows_instead_of_raising(self):
        good = _inventory_row("sku-a", 10)
        bad = FailedRow(row=["bad"], error=ValueError(), context="bad row")

        inventory, failed = build_inventory(_FakeReader([good, bad]))

        assert list(inventory) == ["sku-a"]
        assert failed == [bad]

    def test_raises_datasourceerror_when_no_rows_survive(self):
        bad = FailedRow(row=["bad"], error=ValueError(), context="bad row")

        with pytest.raises(DataSourceError):
            build_inventory(_FakeReader([bad]))


class TestAllocateLandedCosts:
    def test_allocates_cost_to_matching_inventory_row(self):
        inv_row = _inventory_row("sku-a", 10)
        inventory = {"sku-a": inv_row}
        cost_row = _landed_cost_row("sku-a", qty=10, unit_cost=2)

        failed = allocate_landed_costs(_FakeReader([cost_row]), inventory)

        assert failed == []
        assert inv_row.unallocated == Decimal(0)
        assert inv_row.total_cost == Decimal(20)

    def test_purchase_for_sku_not_in_inventory_is_recorded_as_failed(self):
        cost_row = _landed_cost_row("missing-sku", qty=10, unit_cost=2)

        failed = allocate_landed_costs(_FakeReader([cost_row]), {})

        assert len(failed) == 1
        assert failed[0].row is cost_row
        assert failed[0].context == "Purchase for item that is not in inventory."

    def test_collects_failedrow_from_reader_instead_of_raising(self):
        bad = FailedRow(row=["bad"], error=ValueError(), context="bad row")

        failed = allocate_landed_costs(_FakeReader([bad]), {})

        assert failed == [bad]
