from io import BytesIO
from unittest import TestCase
from unittest.mock import patch

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from inventory_kits.reader import DataSourceError, ExcelKitReader

# consts from invetnroy_kits.reader for ease of use
_HEADER_ROWS = {0}
_PARENT_SKU_COL = 0
_COMPONENT_SKU_COL = 1
_KIT_QTY_COL = 2
_UNIT_COST_COL = 11


class TestKitReader(TestCase):
    @staticmethod
    def get_test_excel_bytes():

        wb = openpyxl.Workbook(write_only=False)
        ws = wb.active
        header = [None] * 20
        for i in range(len(header)):
            header[i] = "header row"  # type: ignore
        ws.append(header)

        row1 = [None] * 20
        row1[_PARENT_SKU_COL] = "parent_sku"  # type: ignore
        row1[_COMPONENT_SKU_COL] = "component"  # type: ignore
        row1[_KIT_QTY_COL] = 2  # type: ignore
        row1[_UNIT_COST_COL] = 2.3123  # type: ignore
        ws.append(row1)

        row2 = [None] * 20
        row2[_PARENT_SKU_COL] = "parent_sku"  # type: ignore
        row2[_COMPONENT_SKU_COL] = "component2"  # type: ignore
        row2[_KIT_QTY_COL] = 1  # type: ignore
        row2[_UNIT_COST_COL] = 15  # type: ignore
        ws.append(row2)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_reader_initialization(self):
        """
        Make sure we get a Workbook, worksheet from ExcelKitReader __init__ :)
        """
        real_wb = openpyxl.load_workbook(
            self.get_test_excel_bytes(), read_only=True, data_only=True
        )

        with (
            patch("inventory_kits.reader.xl.load_workbook") as mock_load,
            patch(
                "inventory_kits.reader.ExcelKitReader._save_kits_to_disk"
            ) as mock_save,
        ):
            mock_load.return_value = real_wb

            x = ExcelKitReader("fake_file.xslx")
            assert isinstance(x.wb, Workbook)
            assert isinstance(x.data_source, ReadOnlyWorksheet)
            x.close(save_new=True)
            assert mock_save.called

    def test_reader_initialization_complains_about_multiple_sheets(self):
        def _helper():
            """Return wb with 2 Worksheets."""
            wb = openpyxl.Workbook(write_only=False)
            wb.create_sheet()
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        real_wb = openpyxl.load_workbook(_helper(), read_only=True, data_only=True)
        with (
            patch("inventory_kits.reader.xl.load_workbook") as mock_load,
        ):
            mock_load.return_value = real_wb

            with self.assertRaises(DataSourceError):
                ExcelKitReader("fake_file.xslx")

    def test_reader_raises_when_row_data_incomplete(self):
        real_wb = openpyxl.load_workbook(
            self.get_test_excel_bytes(), read_only=True, data_only=True
        )

        with (
            patch("inventory_kits.reader.xl.load_workbook") as mock_load,
            patch(
                "inventory_kits.reader.ExcelKitReader._save_kits_to_disk"
            ) as mock_save,
        ):
            mock_load.return_value = real_wb

            x = ExcelKitReader("fake_file.xslx")
            x.process_sellercloud_kit_export()

            assert mock_load.called
            assert mock_save.called
            assert "parent_sku" in x.kits
            assert (
                "component" in x.kits["parent_sku"]
                and "component2" in x.kits["parent_sku"]
            )

    # def test_reader_
