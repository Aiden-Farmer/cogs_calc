from __future__ import annotations

import sys
from datetime import datetime
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.data.datarows import Header, InventoryRow
from src.data.excel.reader import (
    CouldNotOpenFile,
    ExcelDataSource,
    ExcelFileReader,
    UnsupportedPlatformError,
    _backup_path,
    _prune_sheet,
    _require_windows,
    remove_wb_dates_after_target,
)
from src.data.reader import AbstractReader

_INV_HEADER = Header.inventory_row(sku=0, base_sku=1, inventory=2)
_OLE_FILE_SIG = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _write_workbook(tmp_path, rows, title="Sheet") -> str:
    wb = openpyxl.Workbook()
    wb.active.title = title
    for row in rows:
        wb.active.append(row)
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return str(path)


class TestInitializeData:
    def test_raises_typeerror_for_chartsheet(self, tmp_path):
        path = _write_workbook(tmp_path, [["header"]], title="Chart1")
        chart_wb = MagicMock()
        chart_wb.__getitem__.return_value = object()  # not a ReadOnlyWorksheet

        with patch("src.data.excel.reader.xl.load_workbook", return_value=chart_wb):
            with pytest.raises(TypeError):
                ExcelFileReader(
                    ExcelDataSource(path, "Chart1"), _INV_HEADER, InventoryRow
                )

    def test_password_protected_file_is_decrypted_and_opened(self, tmp_path):
        encrypted_path = tmp_path / "protected.xlsx"
        encrypted_path.write_bytes(_OLE_FILE_SIG + b"\x00" * 8)

        real_wb = openpyxl.load_workbook(
            _write_workbook(tmp_path, [["header"]]), read_only=True, data_only=True
        )
        mock_office_file = MagicMock()

        with (
            patch("src.data.excel.reader.xl.load_workbook", return_value=real_wb),
            patch("src.data.excel.reader.getpass", return_value="secret"),
            patch(
                "src.data.excel.reader.msoffcrypto.OfficeFile",
                return_value=mock_office_file,
            ),
        ):
            reader = ExcelFileReader(
                ExcelDataSource(str(encrypted_path), "Sheet"), _INV_HEADER, InventoryRow
            )

        mock_office_file.load_key.assert_called_once_with(password="secret")
        mock_office_file.decrypt.assert_called_once()
        assert reader.wb is real_wb

    def test_raises_couldnotopenfile_when_decryption_fails(self, tmp_path):
        encrypted_path = tmp_path / "protected.xlsx"
        encrypted_path.write_bytes(_OLE_FILE_SIG + b"\x00" * 8)

        with (
            patch.object(
                ExcelFileReader, "_handle_password_protected_xl", return_value=None
            ),
            pytest.raises(CouldNotOpenFile),
        ):
            ExcelFileReader(
                ExcelDataSource(str(encrypted_path), "Sheet"),
                _INV_HEADER,
                InventoryRow,
            )

    def test_raises_valueerror_when_initialize_data_returns_falsy(self):
        with (
            patch.object(
                ExcelFileReader, "_initialize_data", return_value=(None, None)
            ),
            pytest.raises(ValueError),
        ):
            ExcelFileReader(
                ExcelDataSource("book.xlsx", "Sheet"), _INV_HEADER, InventoryRow
            )


class TestExcelFileReaderMisc:
    def test_close_closes_workbook(self, tmp_path):
        path = _write_workbook(tmp_path, [["header"]])
        reader = ExcelFileReader(
            ExcelDataSource(path, "Sheet"), _INV_HEADER, InventoryRow
        )

        with patch.object(reader.wb, "close") as mock_close:
            reader.close()

        mock_close.assert_called_once()

    def test_iter_raw_skips_header_row(self, tmp_path):
        path = _write_workbook(tmp_path, [["header"], ["a", "a", 1], ["b", "b", 2]])
        reader = ExcelFileReader(
            ExcelDataSource(path, "Sheet"), _INV_HEADER, InventoryRow
        )

        rows = list(reader._iter_raw())

        assert rows == [("a", "a", 1), ("b", "b", 2)]

    def test_readline_parses_rows_into_return_type(self, tmp_path):
        path = _write_workbook(
            tmp_path, [["header"], ["a", "a", 1], ["bad-sku", "mismatch", 2]]
        )
        reader = ExcelFileReader(
            ExcelDataSource(path, "Sheet"), _INV_HEADER, InventoryRow
        )

        # Bypass the @split_kits decorator (it depends on process-wide kit data
        # loaded from disk at import time) to exercise the base read pipeline directly.
        rows = list(AbstractReader.readline(reader))

        assert len(rows) == 1
        assert rows[0].sku == "a"


class TestRequireWindows:
    def test_raises_on_non_windows(self, monkeypatch):
        monkeypatch.setattr(sys, "platform", "linux")

        with pytest.raises(UnsupportedPlatformError):
            _require_windows("some feature")

    def test_no_op_on_windows(self, monkeypatch):
        monkeypatch.setattr(sys, "platform", "win32")

        _require_windows("some feature")  # should not raise


class TestBackupPath:
    def test_backup_path_preserves_dir_and_suffix_and_is_unique_looking(
        self, tmp_path
    ):
        original = tmp_path / "inventory.xlsx"

        backup = _backup_path(str(original))

        assert backup.parent == original.parent
        assert backup.suffix == ".xlsx"
        assert backup.name.startswith("inventory_backup_")
        assert backup != original


class TestPruneSheet:
    def _make_sheet(self, values: dict[int, object], last_row: int):
        deleted: list[int] = []
        sheet = MagicMock()
        sheet.UsedRange.Row = 1
        sheet.UsedRange.Rows.Count = last_row
        sheet.Cells.side_effect = lambda row, col: SimpleNamespace(
            Value=values.get(row)
        )
        sheet.Rows.side_effect = lambda row: SimpleNamespace(
            Delete=lambda: deleted.append(row)
        )
        sheet.deleted = deleted
        return sheet

    def test_deletes_rows_with_blank_or_after_target_date_keeps_the_rest(self):
        target = datetime(2024, 1, 1)
        values = {
            2: datetime(2023, 6, 1),  # before target -> kept
            3: None,  # blank -> deleted
            4: datetime(2024, 6, 1),  # after target -> deleted
            5: datetime(2024, 1, 1),  # equal to target -> kept
        }
        sheet = self._make_sheet(values, last_row=5)

        _prune_sheet(sheet, date_col=0, target=target)

        assert sorted(sheet.deleted) == [3, 4]

    def test_skips_header_row(self):
        sheet = self._make_sheet({1: None}, last_row=1)

        _prune_sheet(sheet, date_col=0, target=datetime(2024, 1, 1))

        assert sheet.deleted == []

    def test_converts_0_indexed_date_col_to_excels_1_indexed_column(self):
        sheet = self._make_sheet({2: None}, last_row=2)

        _prune_sheet(sheet, date_col=4, target=datetime(2024, 1, 1))

        sheet.Cells.assert_called_once_with(2, 5)


class TestRemoveWbDatesAfterTarget:
    def _fake_win32_module(self, mock_app):
        fake_module = MagicMock()
        fake_module.Dispatch.return_value = mock_app
        return fake_module

    def test_raises_on_non_windows_before_touching_win32com(
        self, tmp_path, monkeypatch
    ):
        monkeypatch.setattr(sys, "platform", "linux")

        with pytest.raises(UnsupportedPlatformError):
            remove_wb_dates_after_target(
                str(tmp_path / "inventory.xlsx"), datetime(2024, 1, 1), {}
            )

    def test_backs_up_prunes_recalculates_and_saves_on_success(
        self, tmp_path, monkeypatch
    ):
        monkeypatch.setattr(sys, "platform", "win32")
        src_path = tmp_path / "inventory.xlsx"
        src_path.write_bytes(b"pretend workbook bytes")

        mock_app = MagicMock()
        mock_wb = mock_app.Workbooks.Open.return_value
        mock_sheet = mock_wb.Sheets.return_value
        mock_sheet.UsedRange.Row = 1
        mock_sheet.UsedRange.Rows.Count = 1  # header only, nothing to prune
        monkeypatch.setitem(
            sys.modules, "win32com.client", self._fake_win32_module(mock_app)
        )

        remove_wb_dates_after_target(
            str(src_path), datetime(2024, 1, 1), {"Transactions": 0}
        )

        backups = list(tmp_path.glob("inventory_backup_*.xlsx"))
        assert len(backups) == 1
        assert backups[0].read_bytes() == src_path.read_bytes()

        mock_wb.Sheets.assert_called_once_with("Transactions")
        mock_app.CalculateFullRebuild.assert_called_once()
        mock_wb.Save.assert_called_once()
        mock_wb.Close.assert_called_once_with(SaveChanges=False)
        mock_app.Quit.assert_called_once()

    def test_discards_edits_and_still_quits_on_failure(self, tmp_path, monkeypatch):
        monkeypatch.setattr(sys, "platform", "win32")
        src_path = tmp_path / "inventory.xlsx"
        src_path.write_bytes(b"pretend workbook bytes")

        mock_app = MagicMock()
        mock_wb = mock_app.Workbooks.Open.return_value
        mock_wb.Sheets.side_effect = RuntimeError("boom")
        monkeypatch.setitem(
            sys.modules, "win32com.client", self._fake_win32_module(mock_app)
        )

        with pytest.raises(RuntimeError):
            remove_wb_dates_after_target(
                str(src_path), datetime(2024, 1, 1), {"Transactions": 0}
            )

        mock_wb.Save.assert_not_called()
        mock_wb.Close.assert_called_once_with(SaveChanges=False)
        mock_app.Quit.assert_called_once()
