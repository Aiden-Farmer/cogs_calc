from __future__ import annotations

import io
from collections.abc import Generator
from dataclasses import dataclass
from getpass import getpass
from typing import TypeVar
from zipfile import BadZipFile
from datetime import datetime

import msoffcrypto
import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from typing import Iterable

from ..datarows import RowLike
from ..reader import AbstractReader
from ..reader import Header
from ..datarow_mutation_utils import split_kits

T = TypeVar("T", bound="RowLike")


@dataclass
class ExcelDataSource:
    wb_path: str
    ws_name: str


class ExcelFileReader(AbstractReader[T, ExcelDataSource]):
    def __init__(
        self, data_source: ExcelDataSource, header: Header, return_type: type[T]
    ):
        self.wb, self.data_source = self._initialize_data(data_source)
        self.rt = return_type
        self.header: Header = header

    def readline(self) -> Generator[T, None, None]:
        return super().readline()

    @split_kits
    def _iter_raw(self) -> Iterable[RowLike]:
        for i, raw in enumerate(self.data_source.iter_rows(values_only=True)):
            if i == 0:
                continue
            yield raw

    def _initialize_data(
        self, ds: ExcelDataSource
    ) -> tuple[Workbook, ReadOnlyWorksheet]:
        try:
            wb = xl.load_workbook(
                filename=ds.wb_path,
                read_only=True,
                data_only=True,
            )
        except BadZipFile as e:
            decrypted = self._handle_password_protected_xl(filename=ds.wb_path)
            if decrypted is None:
                raise CouldNotOpenFile(
                    f"{ds.wb_path} is either not a valid xlsx or could not be decrypted",
                ) from e
            wb = xl.load_workbook(
                filename=decrypted,
                read_only=True,
                data_only=True,
            )

        ws = wb[ds.ws_name]
        if not isinstance(ws, ReadOnlyWorksheet):
            raise TypeError("Data must be in a Worksheet, not ChartSheet.")
        return wb, ws

    def close(self) -> None:
        """Free wb."""
        self.wb.close()

    @staticmethod
    def _handle_password_protected_xl(filename) -> io.BytesIO | None:
        decrypted = io.BytesIO()
        try:
            with open(filename, "rb") as f:
                # Check for OLE File Hexadecimal signature.
                header = f.read(8)
                if header != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
                    return None

                u_input_pass = getpass(
                    f"{filename} is password protected, please enter password: ",
                )
                file = msoffcrypto.OfficeFile(f)
                file.load_key(password=u_input_pass)
                file.decrypt(decrypted)
                return decrypted
        except BadZipFile:
            return None


def remove_wb_dates_after_target(self, target: datetime):
    # release resource so we can reopen with write permissions.
    raise NotImplementedError

class CouldNotOpenFile(Exception):
    pass
