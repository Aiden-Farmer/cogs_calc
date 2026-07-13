import io
from abc import ABC, abstractmethod
from dataclasses import dataclass
from typing import TypeVar, Type, Generic, Iterator, Iterable, Any, Generator

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from zipfile import BadZipFile
import msoffcrypto
import io
from getpass import getpass

from data.datarows import RowLike 


T = TypeVar("T", bound="RowLike")
DS = TypeVar("DS")

class AbstractReader(ABC, Generic[T, DS]):
    def __init__(self, data_source: DS, return_type: Type[T]):
        self.data = self._initialize_data(data_source)
        self.rt = return_type
        
    
    def readlineo(self) -> Generator[T]:
        rows: Iterable[T|None] = (self.rt.from_row(raw) for raw in self._iter_raw())
        valid_rows: Iterable[T] = (r for r in rows if r is not None)

        if not self.rt.must_sort:
            yield from valid_rows
            return
    
        sort_key = getattr(self.rt, "sort_key", None)
        if not sort_key:
            raise RowLikeConfigError("If must_sort=True, sort_key must be defined on RowLike object.")
        yield from sorted(valid_rows, key=sort_key, reverse=True)
        

    @abstractmethod
    def _initialize_data(self, ds) -> Any:
        ...

    @abstractmethod
    def _iter_raw(self) -> Iterable[T]:
        ...

    def close(self) -> None:
        """Default no-op; override where teardown needed."""
        pass

    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


@dataclass
class ExcelDataSource:
    wb_path: str
    ws_name: str

class ExcelFileReader(AbstractReader[T, ExcelDataSource]):
    def __init__(self, data_source: ExcelDataSource, return_type: Type[T]):
        self.wb, self.data_source = self._initialize_data(data_source)
        self.rt = return_type

    def _iter_raw(self) -> Iterator[Any]:
        for i, raw in enumerate(self.data_source.iter_rows(values_only=True)):
                if i == 0:
                    continue
                yield raw 
    
    def _initialize_data(self, ds: ExcelDataSource) -> tuple[Workbook, ReadOnlyWorksheet]:
        try:
            wb = xl.load_workbook(filename=ds.wb_path, read_only=True, data_only=True)
        except BadZipFile as e:
            decrypted = self._handle_password_protected_xl(filename=ds.wb_path)
            if decrypted is None:
                raise CouldNotOpenFile(f"{ds.wb_path} is either not a valid xlsx or could not be decrypted") from e
            wb = xl.load_workbook(filename=decrypted, read_only=True, data_only=True)

        ws = wb[ds.ws_name]
        if not isinstance(ws, ReadOnlyWorksheet):
            raise TypeError("Data must be in a Worksheet, not ChartSheet.")
        return wb, ws
    
    def close(self) -> None:
        self.wb.close()
            
 
    @staticmethod
    def _handle_password_protected_xl(filename) -> io.BytesIO|None:
        decrypted = io.BytesIO()
        try:
            with open(filename, 'rb') as f:
                # Check for OLE File Hexadecimal signature.
                header = f.read(8)
                if header != b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                    return None
                
                u_input_pass = getpass(f"{filename} is password protected, please enter password: ")
                file = msoffcrypto.OfficeFile(f)
                file.load_key(password=u_input_pass)
                file.decrypt(decrypted)
                return decrypted
        except BadZipFile:
            return None





class CouldNotOpenFile(BaseException):
    pass

class RowLikeConfigError(BaseException):
    pass