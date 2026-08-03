from typing import Iterable

from datetime import datetime
import openpyxl as xl
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from data import DataSourceError


class InventoryTransfer:
    from_sku: str
    to_sku: str
    qty: str
    date: datetime


_HEADER_ROWS = {0}
_FROM_SKU_COL = 0
_TO_SKU_COL = 1
_TRANSFER_QTY = 2
_TRANSFER_DATE = 3


class TransferFileReader:
    def __init__(self, filename):
        self.wb, self.data_source = self._initialize_data(filename)

    def _iter(self, h_rows=_HEADER_ROWS) -> Iterable[tuple]:
        """Needless wrapper func?"""
        for i, raw in enumerate(self.data_source.iter_rows(values_only=True)):
            if i in h_rows:
                continue
            yield raw

    def _initialize_data(self, filename) -> tuple[Workbook, ReadOnlyWorksheet]:
        wb = xl.load_workbook(
            filename=filename,
            read_only=True,
            data_only=True,
        )

        if len(wb.sheetnames) > 1:
            raise DataSourceError("Transfers File must have exactly one sheet.")

        ws = wb.active
        if not isinstance(ws, ReadOnlyWorksheet):
            raise TypeError(f"Data must be in Worksgheet, not {type(ws)}")

        return wb, ws

    def close(self, save_new=False) -> None:
        """Free wb"""
        if save_new:
            ...
        self.wb.close()
