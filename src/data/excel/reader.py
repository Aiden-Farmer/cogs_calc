from __future__ import annotations

import io
import shutil
import sys
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from datetime import UTC, datetime
from getpass import getpass
from pathlib import Path
from typing import TypeVar
from warnings import filterwarnings
from zipfile import BadZipFile

import msoffcrypto
import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from ..datarow_mutation_utils import split_kits
from ..datarows import RowLike
from ..reader import AbstractReader, FailedRow, Header

T = TypeVar("T", bound="RowLike")

filterwarnings("ignore", category=UserWarning, module="openpyxl")

_OLE_FILE_SIG = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


@dataclass
class ExcelDataSource:
    wb_path: str
    ws_name: str


class ExcelFileReader(AbstractReader[T, ExcelDataSource]):
    def __init__(
        self, data_source: ExcelDataSource, header: Header, return_type: type[T]
    ):
        self.wb, self.data_source = self._initialize_data(data_source)
        self.rt = return_type
        self.header: Header = header

        if not self.wb or not self.data_source:
            raise ValueError

    @split_kits
    def readline(self) -> Iterator[T | FailedRow]:
        return super().readline()

    def _iter_raw(self) -> Iterable[tuple]:
        for i, raw in enumerate(self.data_source.iter_rows(values_only=True)):
            if i == 0:
                continue
            yield raw

    def _initialize_data(
        self, ds: ExcelDataSource
    ) -> tuple[Workbook, ReadOnlyWorksheet]:
        with open(ds.wb_path, "rb") as f:
            header = f.read(8)

        if header == _OLE_FILE_SIG:
            io_stream = self._handle_password_protected_xl(ds.wb_path)
            if not io_stream:
                raise CouldNotOpenFile(
                    f"{ds.wb_path} is either not a valid xlsx or could not be decrypted",
                )
            wb = xl.load_workbook(
                filename=io_stream,
                read_only=True,
                data_only=True,
            )
        else:
            wb = xl.load_workbook(
                filename=ds.wb_path,
                read_only=True,
                data_only=True,
            )

        ws = wb[ds.ws_name]
        if not isinstance(ws, ReadOnlyWorksheet):
            raise TypeError("Data must be in a Worksheet, not ChartSheet.")
        return wb, ws

    def close(self) -> None:
        """Free wb."""
        self.wb.close()

    @staticmethod
    def _handle_password_protected_xl(filename) -> io.BytesIO | None:
        decrypted = io.BytesIO()
        try:
            with open(filename, "rb") as f:
                # Check for OLE File Hexadecimal signature.
                header = f.read(8)
                if header != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
                    return None

                u_input_pass = getpass(
                    f"{filename} is password protected, please enter password: ",
                )
                file = msoffcrypto.OfficeFile(f)
                file.load_key(password=u_input_pass)
                file.decrypt(decrypted)
                return decrypted
        except BadZipFile:
            return None


class UnsupportedPlatformError(OSError):
    pass


def _require_windows(feature: str) -> None:
    """Raise clearly instead of letting a win32-only call fail lower down."""
    if sys.platform != "win32":
        raise UnsupportedPlatformError(
            f"{feature} requires Windows (uses Excel COM automation via pywin32)."
        )


# Maps sheet name -> 0-indexed column holding that row's date. Identifies
# which sheets in a workbook are "transaction" sheets for
# remove_wb_dates_after_target, and where to find the date on each.
type TransactionSheetDateColumns = dict[str, int]


def remove_wb_dates_after_target(
    wb_path: str,
    target: datetime,
    transaction_sheets: TransactionSheetDateColumns,
) -> None:
    """
    Delete rows missing a date, or dated after `target`, from each sheet
    named in `transaction_sheets`, then recalculate and save.

    Uses Excel COM automation (pywin32) rather than openpyxl: openpyxl has
    no formula engine, so it can't recalculate formulas that reference the
    deleted rows. The original file is backed up before anything is
    changed; on any failure the in-memory edits are discarded and the
    original file on disk is left untouched.
    """
    _require_windows("Removing out-of-range transaction rows")

    import win32com.client as win32

    shutil.copy2(wb_path, _backup_path(wb_path))

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(Path(wb_path).resolve()))
        for sheet_name, date_col in transaction_sheets.items():
            _prune_sheet(wb.Sheets(sheet_name), date_col, target)
        excel.CalculateFullRebuild()
        wb.Save()
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()


def _backup_path(wb_path: str) -> Path:
    src = Path(wb_path)
    stamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    return src.with_name(f"{src.stem}_backup_{stamp}{src.suffix}")


def _prune_sheet(sheet, date_col: int, target: datetime) -> None:
    """Delete rows (bottom-up, skipping header row 1) whose date_col cell
    is blank or holds a date after target. Bottom-up avoids row indices
    shifting out from under us as we delete."""
    excel_col = date_col + 1  # Header-style 0-index -> Excel's 1-index
    used = sheet.UsedRange
    last_row = used.Row + used.Rows.Count - 1
    for row in range(last_row, 1, -1):
        value = sheet.Cells(row, excel_col).Value
        if value is None or (isinstance(value, datetime) and value > target):
            sheet.Rows(row).Delete()


class CouldNotOpenFile(Exception):
    pass
