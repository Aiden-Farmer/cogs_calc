from collections import defaultdict
from collections.abc import Iterable
from decimal import Decimal, InvalidOperation
from pickle import dump, load

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from src.data import DataSourceError
from src.data.reader import AbstractReader

# Row locations, specific to sellercloud's standard kit export file format
# so I'm not bothering with an easily user-changeable config.
#  If sellercloud changes the file format, it'll be obvious when DataSourceError is raised.

_HEADER_ROWS = {0}
_PARENT_SKU_COL = 0
_COMPONENT_SKU_COL = 1
_KIT_QTY_COL = 2
_UNIT_COST_COL = 11


class ExcelKitReader(AbstractReader):
    def __init__(self, filename: str):
        self.wb, self.data_source = self._initialize_data(filename)
        self.kits: dict[str, dict[str, dict]] = defaultdict(dict)
        self.total_costs: dict[str, Decimal] = {}

        if not self.wb or not self.data_source:
            raise ValueError

    def readline(self):
        raise NotImplementedError(
            "ExcelKitReader does not produce RowLike rows; "
            "read_kits_from-export is does not produce consumable rows."
            "To read rows, access ExcelKitReader.kits after calling "
            "self.read_kits_from_export() or self.process_sellercloud_kit_export()"
        )

    def process_sellercloud_kit_export(self):
        """Read a sellercloud kit export and save to disk"""
        self.read_kits_from_export()
        self.update_avco()  # Used to update average cost of kit parent.
        self._save_kits_to_disk()
        self._save_kits_total_cost_to_disk()

    def read_kits_from_export(self) -> None:
        """Reads sellercloud formatted kit file to self.kits"""

        for row in self._iter_raw():
            try:
                parent_sku = row[_PARENT_SKU_COL]
                child_sku = row[_COMPONENT_SKU_COL]
                kit_qty = int(row[_KIT_QTY_COL])
                unit_cost = Decimal(row[_UNIT_COST_COL])
                self.kits[parent_sku][child_sku] = {"qty": kit_qty, "cost": unit_cost}
            except (ValueError, InvalidOperation) as e:
                raise DataSourceError(
                    f"{row} is invalid. Check Sellercloud Kit export and make sure columns have not changed."
                ) from e

    def update_avco(self):
        """Second pass of kit keys in self.kits after reading to self.kits
        to build cost ratio of kit component as a part of a whole.
        RCH cannot trust kit parent cost, so we'll rebuild kit parent cost from total kit component cost ."""
        if not self.kits:
            raise ValueError("Missing kit information")

        for kit, components in self.kits.items():
            total_cost = Decimal(0)
            for component in components:
                total_cost += (
                    components[component]["qty"] * components[component]["cost"]
                )

            self.total_costs[kit] = total_cost
            if not total_cost:
                print(f"no cost ratio from sellerlcoud cloud... {kit}")
                continue

            for component in components:
                self.kits[kit][component]["pc_of_total_cost"] = (
                    components[component]["qty"]
                    * components[component]["cost"]
                    / total_cost
                )

    def _load_kits_from_disk(self) -> None:
        with open("kits.obj", "rb") as f:
            self.kits = load(f)

    def _save_kits_to_disk(self) -> None:
        with open("kits.obj", "wb") as f:
            dump(self.kits, f)

    def _save_kits_total_cost_to_disk(self) -> None:
        with open("costs.obj", "wb") as f:
            dump(self.total_costs, f)

    def _iter_raw(self, h_rows=_HEADER_ROWS) -> Iterable[tuple]:
        """yields rows from datasource, skips first row assumming header"""
        for i, raw in enumerate(self.data_source.iter_rows(values_only=True)):
            if i in h_rows:
                continue
            yield raw

    def _initialize_data(self, ds) -> tuple[Workbook, ReadOnlyWorksheet]:

        wb = xl.load_workbook(
            filename=ds,
            read_only=True,
            data_only=True,
        )

        if len(wb.sheetnames) > 1:
            raise DataSourceError("Must have exactly one Worksheet in file.")

        ws = wb.active
        if not isinstance(ws, ReadOnlyWorksheet):
            raise TypeError(f"Data must be in a Worksheet, not {type(ws)}.")

        return wb, ws

    def close(self, save_new=False) -> None:
        """Free wb."""
        if save_new:
            self._save_kits_to_disk()
            self._save_kits_total_cost_to_disk()
        self.wb.close()
